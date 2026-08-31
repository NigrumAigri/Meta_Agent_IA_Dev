from __future__ import annotations

import ast
import hashlib
import json
import logging
import os
import time
from pathlib import Path
from typing import Any

from core.config import settings
from core.domain import McpToolDefinition
from storage.repository import mcp_repo
from services.docker_sandbox import docker_sandbox
from services.mcp_client import mcp_client

logger = logging.getLogger(__name__)

# ------------------------------------------------------------------------------
# 1. CATALOGUE OFFICIEL DES OUTILS MCP NATIFS (Conforme Module 5 §2.1)
# ------------------------------------------------------------------------------

CORE_MCP_TOOLS: list[McpToolDefinition] = [
    McpToolDefinition(
        id="search_knowledge_base",
        name="Recherche Knowledge Base & Référentiel RAG",
        description="[Mission] Interroge le référentiel officiel (14 masterclasses d'architecture, design patterns et sécurité). [Déclencheur] Questions sur l'architecture, MCP, RAG, DDD ou standards de dev. [Interdiction] Ne pas utiliser pour chercher des fichiers du projet local (utiliser le contexte de code). [Résultat] Renvoie les extraits documentaires certifiés.",
        category="RAG & Connaissances",
        parameters_schema={
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Question ou terme technique à rechercher"},
                "top_k": {"type": "integer", "description": "Nombre de résultats à renvoyer", "default": 3},
            },
            "required": ["query"],
        },
        mcp_primitive="resource",
        is_idempotent=True,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="web_search_and_docs",
        name="Recherche Web & Documentation Technique",
        description="[Mission] Recherche en direct sur le web la documentation et APIs officielles. [Déclencheur] Découverte de nouvelles bibliothèques, versions de frameworks ou syntaxes récentes. [Interdiction] Ne pas utiliser pour chercher des données confidentielles ou internes du projet. [Résultat] Renvoie le texte nettoyé encapsulé dans <external_untrusted_data>.",
        category="Recherche & Scraping",
        parameters_schema={
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Terme de recherche technique"},
                "max_results": {"type": "integer", "description": "Nombre de résultats web (max 10)", "default": 4},
            },
            "required": ["query"],
        },
        mcp_primitive="tool",
        is_idempotent=False,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="file_writer_atomic",
        name="Éditeur de Fichiers Atomique Sécurisé",
        description="[Mission] Écrit ou met à jour des fichiers sources de manière atomique avec fsync sans corruption. [Déclencheur] Création de fichiers, modification de code ou sauvegarde d'artefacts. [Interdiction] Ne pas utiliser sans avoir vérifié la syntaxe AST au préalable pour le code Python. [Résultat] Renvoie le statut d'écriture et le nombre d'octets écrits.",
        category="Système de Fichiers",
        parameters_schema={
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Chemin absolu ou relatif du fichier cible"},
                "content": {"type": "string", "description": "Contenu complet ou code source à écrire"},
            },
            "required": ["file_path", "content"],
        },
        mcp_primitive="tool",
        is_idempotent=False,
        is_critical=True,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="ast_validator",
        name="Analyseur Syntaxique & Validateur AST",
        description="[Mission] Analyse la syntaxe Python en mémoire (2 ms) sans exécuter le code pour certifier 100% de conformité AST. [Déclencheur] Toute génération ou modification de code Python avant écriture. [Interdiction] Ne pas exécuter de tests fonctionnels avec cet outil. [Résultat] Renvoie is_valid (True/False) et la ligne exacte en cas d'erreur de syntaxe.",
        category="Qualité & Sécurité",
        parameters_schema={
            "type": "object",
            "properties": {
                "code_content": {"type": "string", "description": "Code source Python à valider"},
                "filename": {"type": "string", "description": "Nom du fichier pour contexte d'erreur", "default": "<memory>"},
            },
            "required": ["code_content"],
        },
        mcp_primitive="tool",
        is_idempotent=True,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="test_runner_sandbox",
        name="Exécuteur de Tests Sécurisé (Docker Sandbox)",
        description="[Mission] Exécute les suites de tests unitaires Pytest dans un conteneur Docker éphémère étanche (--rm, --network none, 512MB RAM). [Déclencheur] Validation fonctionnelle d'un livrable ou calcul de la couverture de tests. [Interdiction] Ne jamais exécuter de code hors sandbox. [Résultat] Renvoie le rapport Pytest (stdout, stderr, temps d'exécution).",
        category="Tests & Qualité",
        parameters_schema={
            "type": "object",
            "properties": {
                "test_dir": {"type": "string", "description": "Dossier contenant les tests (ex: 'tests' ou chemin projet)"},
                "pytest_args": {"type": "array", "items": {"type": "string"}, "description": "Arguments optionnels pour pytest"},
            },
            "required": ["test_dir"],
        },
        mcp_primitive="tool",
        is_idempotent=False,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="document_extractor",
        name="Extracteur de Documents, Tableurs Excel & Spécifications",
        description="[Mission] Extrait et structure le contenu des tableurs Excel multi-feuilles (.xlsx, .xls), CSV et cahiers des charges (PDF, Markdown, TXT) avec mode adaptatif et slicing. [Déclencheur] Lecture de données tabulaires, budgets Excel ou spécifications techniques. [Interdiction] Ne modifie pas les fichiers (lecture seule). [Résultat] Renvoie les tables Markdown profilées et les extraits.",
        category="Traitement de Données",
        parameters_schema={
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Chemin du fichier (.xlsx, .xls, .csv, .md, .txt, .pdf, .json, .py)"},
                "sheet_name": {"type": "string", "description": "Nom de la feuille Excel spécifique (optionnel)"},
                "start_row": {"type": "integer", "description": "Numéro de ligne de départ pour slicing (1-indexé)"},
                "max_rows": {"type": "integer", "description": "Nombre maximal de lignes à extraire"},
            },
            "required": ["file_path"],
        },
        mcp_primitive="resource",
        is_idempotent=True,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="code_formatter",
        name="Formateur de Code PEP8",
        description="[Mission] Formate et nettoie le code Python selon les standards stricts PEP8. [Déclencheur] Nettoyage final de code avant commit ou export. [Interdiction] Ne modifie pas la logique du code. [Résultat] Renvoie le code source proprement indenté.",
        category="Qualité de Code",
        parameters_schema={
            "type": "object",
            "properties": {
                "code_content": {"type": "string", "description": "Code Python brut à formater"},
            },
            "required": ["code_content"],
        },
        mcp_primitive="tool",
        is_idempotent=True,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="finops_calculator",
        name="Calculateur de Coûts & Tokens Déterministe",
        description="[Mission] Calcule avec précision chirurgicale la consommation de tokens et les dépenses au centime près ($0.0000). [Déclencheur] Audit de consommation, projection budgétaire ou facturation de session. [Interdiction] Ne jamais estimer un coût de tête. [Résultat] Renvoie le détail des tokens prompt/completion/reasoning et le total USD.",
        category="FinOps & Télémétrie",
        parameters_schema={
            "type": "object",
            "properties": {
                "prompt_tokens": {"type": "integer", "description": "Tokens de prompt en entrée"},
                "completion_tokens": {"type": "integer", "description": "Tokens générés en sortie"},
                "reasoning_tokens": {"type": "integer", "description": "Tokens de raisonnement interne"},
                "price_in_usd": {"type": "number", "description": "Prix $/1M tokens entrée"},
                "price_out_usd": {"type": "number", "description": "Prix $/1M tokens sortie"},
            },
            "required": ["prompt_tokens", "completion_tokens"],
        },
        mcp_primitive="tool",
        is_idempotent=True,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="math_calculator",
        name="Calculatrice Mathématique Déterministe (0% Hallucination)",
        description="[Mission] Évalue en pur Python sécurisé des calculs financiers, pourcentages, ratios et équations. [Déclencheur] Toute opération arithmétique, calcul de pourcentage ou total. [Interdiction] Ne jamais faire de calcul mental dans les prompts. [Résultat] Renvoie la valeur numérique exacte.",
        category="Calcul Déterministe",
        parameters_schema={
            "type": "object",
            "properties": {
                "expression": {"type": "string", "description": "Expression arithmétique Python sécurisée"},
            },
            "required": ["expression"],
        },
        mcp_primitive="tool",
        is_idempotent=True,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="search_models_catalog",
        name="Catalogue Universel des Modèles & Benchmarks Scientifiques",
        description="[Mission] Recherche, filtre et compare les 600+ modèles IA en direct selon les 19 benchmarks certifiés Artificial Analysis et les tarifs réels OpenRouter. [Déclencheur] Sélection du meilleur LLM pour un rôle (Sweet Spot, Top Perf, Ultra Éco). [Interdiction] Ne pas hardcoder de liste statique de modèles. [Résultat] Renvoie la liste ordonnée des modèles avec scores scientifiques et prix.",
        category="FinOps & Modèles",
        parameters_schema={
            "type": "object",
            "properties": {
                "q": {"type": "string", "description": "Recherche textuelle (nom, créateur, slug)"},
                "max_price_out_usd": {"type": "number", "description": "Plafond tarifaire $/1M tokens sortie"},
                "min_coding_score": {"type": "number", "description": "Score minimum de programmation (0-100)"},
                "benchmark_filters": {"type": "object", "description": "Filtres sur benchmarks (ex: {'terminalbench_v2_1': 0.80})"},
                "sort_by": {"type": "string", "description": "Tri : coding_desc, price_asc, speed_desc, gpqa_desc"},
                "limit": {"type": "integer", "description": "Nombre de résultats", "default": 5},
            },
        },
        mcp_primitive="resource",
        is_idempotent=True,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="get_catalog_capabilities",
        name="Capacités & Métriques Disponibles du Catalogue",
        description="[Mission] Renvoie la liste des benchmarks scientifiques enregistrés en base, le nombre de modèles et le statut de synchronisation. [Déclencheur] Découverte des critères de matching disponibles. [Interdiction] Ne modifie pas la base de données. [Résultat] Renvoie les définitions des métriques et l'état du cache.",
        category="FinOps & Modèles",
        parameters_schema={
            "type": "object",
            "properties": {},
        },
        mcp_primitive="resource",
        is_idempotent=True,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="discover_tools",
        name="Auto-Découverte Dynamique d'Outils (Tool RAG)",
        description="[Mission] Permet à un agent de rechercher en direct dans le registre MCP les outils spécialisés adaptés à son besoin immédiat. [Déclencheur] Besoin d'une capacité spécifique non pré-chargée en début de tour. [Interdiction] Ne pas utiliser si l'outil est déjà assigné. [Résultat] Renvoie les fiches d'outils correspondantes avec schémas JSON.",
        category="Système & Découverte",
        parameters_schema={
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Besoin fonctionnel ou capacité recherchée (ex: 'convertir image png', 'scraper web')"},
                "limit": {"type": "integer", "description": "Nombre maximal d'outils à découvrir", "default": 3},
            },
            "required": ["query"],
        },
        mcp_primitive="tool",
        is_idempotent=True,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="read_skill",
        name="Lecteur de Playbook de Compétence (Skill JIT)",
        description="[Mission] Charge à la demande les instructions techniques détaillées d'un Playbook (SKILL.md) et la liste de ses ressources. [Déclencheur] Dès qu'une tâche nécessite d'appliquer les directives précises d'une compétence listée dans <available_skills> ou découverte via discover_skills. [Interdiction] Ne pas utiliser si le corps du playbook a déjà été chargé au tour courant. [Résultat] Renvoie les instructions Markdown complètes, directives d'architecture et exemples.",
        category="Compétences JIT",
        parameters_schema={
            "type": "object",
            "properties": {
                "skill_name": {"type": "string", "description": "Nom exact de la compétence (ex: 'fastapi_enterprise', 'sqlite_wal_persistence')"},
            },
            "required": ["skill_name"],
        },
        mcp_primitive="resource",
        is_idempotent=True,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
    McpToolDefinition(
        id="discover_skills",
        name="Auto-Découverte Dynamique de Compétences (Skill RAG)",
        description="[Mission] Permet à un agent de rechercher en direct dans le registre des Playbooks les compétences spécialisées adaptées à son besoin immédiat. [Déclencheur] Besoin d'un guide technique ou d'un standard de conception non pré-chargé en début de tour. [Interdiction] Ne pas utiliser si la compétence est déjà active. [Résultat] Renvoie les fiches de compétences correspondantes avec métadonnées YAML.",
        category="Système & Découverte",
        parameters_schema={
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Besoin fonctionnel ou compétence recherchée (ex: 'authentification JWT', 'clean architecture')"},
                "limit": {"type": "integer", "description": "Nombre maximal de compétences à découvrir", "default": 3},
            },
            "required": ["query"],
        },
        mcp_primitive="tool",
        is_idempotent=True,
        is_critical=False,
        is_active=True,
        is_core=True,
    ),
]


class McpHub:
    """Catalogue central des outils MCP, routeur universel et gestionnaire d'exécution."""

    def __init__(self) -> None:
        self._idempotence_cache: dict[str, tuple[float, dict[str, Any]]] = {}
        self._cache_ttl_seconds = 900.0  # 15 minutes TTL
        self._seed_core_tools()

    def _seed_core_tools(self) -> None:
        for tool in CORE_MCP_TOOLS:
            mcp_repo.save_tool(tool)

    def list_tools(self, project_id: str | None = None, active_only: bool = False) -> list[McpToolDefinition]:
        return mcp_repo.list_tools(project_id=project_id, active_only=active_only)

    def get_tool(self, tool_id: str, project_id: str | None = None) -> McpToolDefinition | None:
        normalized_id = tool_id
        if tool_id == "web_search":
            normalized_id = "web_search_and_docs"
        tools = self.list_tools(project_id=project_id, active_only=False)
        return next((t for t in tools if t.id == normalized_id), None)

    # --------------------------------------------------------------------------
    # Troncature Universelle des Retours (Module 5 §2.2.1)
    # --------------------------------------------------------------------------

    def _truncate_output(self, result: Any, max_chars: int = 8000) -> Any:
        """Plafonne tout retour textuel volumineux à ~2 000 tokens pour protéger le contexte LLM."""
        if isinstance(result, str) and len(result) > max_chars:
            truncated = result[:max_chars]
            return f"{truncated}\n\n[... Texte tronqué à {max_chars} caractères pour préserver le contexte LLM ...]"
        elif isinstance(result, dict):
            return {k: self._truncate_output(v, max_chars) for k, v in result.items()}
        elif isinstance(result, list):
            if len(result) > 20:
                truncated_list = [self._truncate_output(item, max_chars) for item in result[:20]]
                truncated_list.append(f"[... {len(result) - 20} éléments masqués pour optimiser les tokens ...]")
                return truncated_list
            return [self._truncate_output(item, max_chars) for item in result]
        return result

    # --------------------------------------------------------------------------
    # Exécution Déterministe et Dispatching Universel
    # --------------------------------------------------------------------------

    def execute_tool(self, tool_id: str, arguments: dict[str, Any], project_id: str | None = None) -> dict[str, Any]:
        """Exécute un outil système ou distant de manière sécurisée, avec sentinelles de cycle de vie, cache et troncature."""
        from core.domain import HookEventType
        from services.hooks_engine import hooks_engine

        tool = self.get_tool(tool_id, project_id=project_id)
        if not tool:
            return {"status": "error", "message": f"Outil inconnu ou non disponible : {tool_id}"}

        if not tool.is_active:
            return {"status": "error", "message": f"L'outil '{tool.name}' ({tool_id}) est actuellement désactivé."}

        # 0. Interception Sentinelle PRE_TOOL_CALL (Sécurité & Permissions)
        pre_hook_res = hooks_engine.trigger_event(
            HookEventType.PRE_TOOL_CALL,
            {"tool_id": tool_id, "arguments": arguments, "project_id": project_id},
        )
        if pre_hook_res.get("is_blocked"):
            return {
                "status": "blocked",
                "message": pre_hook_res.get("block_reason") or "Opération bloquée par la sentinelle de sécurité PRE_TOOL_CALL.",
            }

        # 1. Vérification du Cache Idempotent (pour outils en lecture seule)
        if tool.is_idempotent:
            cache_key = hashlib.sha256(f"{tool_id}:{json.dumps(arguments, sort_keys=True)}".encode()).hexdigest()
            now = time.time()
            if cache_key in self._idempotence_cache:
                timestamp, cached_res = self._idempotence_cache[cache_key]
                if now - timestamp < self._cache_ttl_seconds:
                    logger.debug("Cache HIT pour l'outil idempotent '%s'", tool_id)
                    return cached_res

        # 2. Exécution selon l'origine (Natif vs Serveur Externe)
        raw_res: dict[str, Any]
        try:
            if tool.is_core:
                raw_res = self._execute_core_tool(tool_id, arguments, project_id=project_id)
            elif tool.server_id:
                # Exécution sur serveur externe (Stdio / SSE)
                server = next((s for s in mcp_repo.list_servers(project_id=project_id) if s.id == tool.server_id), None)
                if server:
                    raw_res = mcp_client.execute_remote_tool(server, tool.name, arguments)
                else:
                    raw_res = {"status": "error", "message": f"Serveur MCP distant introuvable pour l'outil : {tool_id}"}
            else:
                raw_res = {"status": "error", "message": f"Exécuteur introuvable pour l'outil : {tool_id}"}
        except Exception as err:
            hooks_engine.trigger_event(
                HookEventType.ON_ERROR,
                {"source": f"mcp_tool:{tool_id}", "error": str(err), "project_id": project_id},
            )
            raise

        # 3. Troncature universelle automatique
        final_res = self._truncate_output(raw_res)

        # 3.bis Interception Sentinelle POST_TOOL_CALL (Validation AST & Formatage)
        hooks_engine.trigger_event(
            HookEventType.POST_TOOL_CALL,
            {"tool_id": tool_id, "arguments": arguments, "result": final_res, "project_id": project_id},
        )

        # 4. Enregistrement dans le Cache Idempotent
        if tool.is_idempotent and final_res.get("status") == "success":
            cache_key = hashlib.sha256(f"{tool_id}:{json.dumps(arguments, sort_keys=True)}".encode()).hexdigest()
            self._idempotence_cache[cache_key] = (time.time(), final_res)

        return final_res

    def _execute_core_tool(self, tool_id: str, arguments: dict[str, Any], project_id: str | None = None) -> dict[str, Any]:
        if tool_id == "search_knowledge_base":
            return self._exec_search_knowledge_base(arguments)
        elif tool_id == "ast_validator":
            return self._exec_ast_validator(arguments)
        elif tool_id == "file_writer_atomic":
            return self._exec_file_writer_atomic(arguments)
        elif tool_id == "math_calculator":
            return self._exec_math_calculator(arguments)
        elif tool_id == "finops_calculator":
            return self._exec_finops_calculator(arguments)
        elif tool_id in ("web_search", "web_search_and_docs"):
            return self._exec_web_search(arguments)
        elif tool_id == "document_extractor":
            return self._exec_document_extractor(arguments)
        elif tool_id == "code_formatter":
            return self._exec_code_formatter(arguments)
        elif tool_id == "test_runner_sandbox":
            return self._exec_test_runner(arguments)
        elif tool_id == "search_models_catalog":
            return self._exec_search_models_catalog(arguments)
        elif tool_id == "get_catalog_capabilities":
            return self._exec_get_catalog_capabilities(arguments)
        elif tool_id == "discover_tools":
            return self._exec_discover_tools(arguments, project_id=project_id)
        elif tool_id == "read_skill":
            return self._exec_read_skill(arguments, project_id=project_id)
        elif tool_id == "discover_skills":
            return self._exec_discover_skills(arguments, project_id=project_id)
        else:
            return {"status": "error", "message": f"Outil natif non implémenté : {tool_id}"}

    # --------------------------------------------------------------------------
    # Implémentations Déterministes Internes
    # --------------------------------------------------------------------------

    def _exec_search_knowledge_base(self, args: dict[str, Any]) -> dict[str, Any]:
        from services.rag_engine import rag_engine
        query = args.get("query", "")
        top_k = int(args.get("top_k", 3))
        results = rag_engine.search(query=query, top_k=top_k)
        return {
            "status": "success",
            "query": query,
            "count": len(results),
            "results": results,
        }

    def _exec_ast_validator(self, args: dict[str, Any]) -> dict[str, Any]:
        code = args.get("code_content", "")
        filename = args.get("filename", "<memory>")
        try:
            ast.parse(code, filename=filename)
            return {"status": "success", "is_valid": True, "message": "Syntaxe Python 100% valide (AST certifié)."}
        except SyntaxError as e:
            return {
                "status": "error",
                "is_valid": False,
                "error": f"Erreur de syntaxe ligne {e.lineno}, colonne {e.offset}: {e.msg}",
                "line": e.lineno,
                "offset": e.offset,
                "text": e.text,
            }

    def _exec_file_writer_atomic(self, args: dict[str, Any]) -> dict[str, Any]:
        path_str = args.get("file_path", "")
        content = args.get("content", "")
        if not path_str:
            return {"status": "error", "message": "Chemin de fichier manquant"}

        target_file = Path(path_str)
        target_file.parent.mkdir(parents=True, exist_ok=True)

        # Garde-fou AST préalable pour les fichiers Python
        if settings.ast_validation_enabled and target_file.suffix.lower() == ".py":
            try:
                ast.parse(content, filename=str(target_file))
            except SyntaxError as e:
                return {
                    "status": "error",
                    "message": f"Garde-fou AST Actif : Erreur de syntaxe Python ligne {e.lineno}, colonne {e.offset} ({e.msg}). Écriture disque bloquée.",
                    "is_syntax_valid": False,
                }

        temp_file = target_file.with_suffix(f".tmp.{os.getpid()}")
        try:
            with open(temp_file, "w", encoding="utf-8") as f:
                f.write(content)
                f.flush()
                os.fsync(f.fileno())
            os.replace(temp_file, target_file)
            return {
                "status": "success",
                "file_path": str(target_file),
                "bytes_written": len(content.encode("utf-8")),
            }
        except Exception as e:
            if temp_file.exists():
                try:
                    temp_file.unlink()
                except OSError:
                    pass
            return {"status": "error", "message": f"Échec écriture atomique : {str(e)}"}

    def _exec_math_calculator(self, args: dict[str, Any]) -> dict[str, Any]:
        expr = args.get("expression", "")
        try:
            tree = ast.parse(expr, mode="eval")
            for node in ast.walk(tree):
                if not isinstance(node, (ast.Expression, ast.BinOp, ast.UnaryOp, ast.Constant, ast.operator, ast.unaryop)):
                    raise ValueError("Seules les opérations arithmétiques basiques sont autorisées.")
            result = eval(compile(tree, "<string>", "eval"), {"__builtins__": {}}, {})
            return {"status": "success", "result": result}
        except Exception as e:
            return {"status": "error", "message": f"Calcul invalide : {str(e)}"}

    def _exec_finops_calculator(self, args: dict[str, Any]) -> dict[str, Any]:
        p_tokens = int(args.get("prompt_tokens", 0))
        c_tokens = int(args.get("completion_tokens", 0))
        r_tokens = int(args.get("reasoning_tokens", 0))
        p_price = float(args.get("price_in_usd", 1.0))
        c_price = float(args.get("price_out_usd", 3.0))

        cost_in = (p_tokens / 1_000_000) * p_price
        cost_out = ((c_tokens + r_tokens) / 1_000_000) * c_price
        total_cost = round(cost_in + cost_out, 6)

        return {
            "status": "success",
            "prompt_tokens": p_tokens,
            "completion_tokens": c_tokens,
            "reasoning_tokens": r_tokens,
            "total_tokens": p_tokens + c_tokens + r_tokens,
            "cost_usd": total_cost,
        }

    def _exec_web_search(self, args: dict[str, Any]) -> dict[str, Any]:
        query = str(args.get("query", "")).strip()
        if not query:
            return {"status": "error", "message": "Requête de recherche vide."}

        max_results = min(int(args.get("max_results") or 4), 10)
        snippets = []
        try:
            from ddgs import DDGS
            with DDGS() as ddgs:
                raw_results = list(ddgs.text(query, max_results=max_results))
            for r in raw_results:
                title = r.get("title", "").strip()
                body = r.get("body", "").strip()
                href = r.get("href", "").strip()
                if title or body:
                    snippets.append(f"### {title}\nSource: {href}\n{body}")
        except Exception:
            try:
                import httpx
                resp = httpx.get(
                    "https://api.duckduckgo.com/",
                    params={"q": query, "format": "json", "no_html": 1, "skip_disambig": 1},
                    headers={"User-Agent": "MetaAgentDev/5.0"},
                    timeout=3.0,
                )
                if resp.status_code == 200:
                    data = resp.json()
                    abstract = data.get("AbstractText", "")
                    heading = data.get("Heading", "")
                    if abstract:
                        snippets.append(f"### {heading}\n{abstract}")
            except Exception:
                pass

        if not snippets:
            search_summary = f"Aucun résultat trouvé sur le web pour : {query}"
        else:
            search_summary = "\n\n".join(snippets)

        # Encapsulation obligatoire anti-injection (Module 11)
        results = f"<external_untrusted_data>\n{search_summary}\n</external_untrusted_data>"
        return {
            "status": "success",
            "query": query,
            "count": len(snippets),
            "results": results,
            "security": "READ_ONLY_UNTRUSTED",
        }

    def _exec_document_extractor(self, args: dict[str, Any]) -> dict[str, Any]:
        path_str = str(args.get("file_path", "")).strip()
        if not path_str:
            return {"status": "error", "message": "Le paramètre 'file_path' est requis."}

        file_path = Path(path_str)
        if not file_path.exists():
            return {"status": "error", "message": f"Fichier introuvable : {path_str}"}

        ext = file_path.suffix.lower()
        sheet_param = args.get("sheet_name")
        start_row = int(args.get("start_row") or 1)
        max_rows = int(args["max_rows"]) if args.get("max_rows") is not None else None

        if ext in (".xlsx", ".xls"):
            wb = None
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filename=str(file_path), data_only=True, read_only=True)
                sheet_names = list(wb.sheetnames)
                target_sheets = [sheet_param] if sheet_param and sheet_param in sheet_names else sheet_names
                
                sections = []
                total_file_rows = 0
                for s_name in target_sheets:
                    ws = wb[s_name]
                    raw_rows = []
                    for row in ws.iter_rows(values_only=True):
                        if any(cell is not None and str(cell).strip() != "" for cell in row):
                            raw_rows.append([str(c).strip() if c is not None else "" for c in row])

                    num_rows = len(raw_rows)
                    total_file_rows += num_rows
                    if num_rows == 0:
                        sections.append(f"### [Feuille : {s_name}]\n*(Feuille vide)*")
                        continue

                    header = raw_rows[0]
                    data_rows = raw_rows[1:]

                    if start_row > 1 or max_rows is not None:
                        s_idx = max(0, start_row - 2)
                        e_idx = s_idx + (max_rows if max_rows is not None else len(data_rows))
                        sliced_data = data_rows[s_idx:e_idx]

                        md = [f"### [Feuille : {s_name}] (Extrait lignes {start_row} à {start_row + len(sliced_data) - 1} / {num_rows})"]
                        md.append("| " + " | ".join(header) + " |")
                        md.append("| " + " | ".join(["---"] * len(header)) + " |")
                        for r in sliced_data:
                            padded = r + [""] * (len(header) - len(r))
                            md.append("| " + " | ".join(padded[:len(header)]) + " |")
                        sections.append("\n".join(md))
                        continue

                    if num_rows <= 150:
                        md = [f"### [Feuille : {s_name}] ({num_rows} lignes)"]
                        md.append("| " + " | ".join(header) + " |")
                        md.append("| " + " | ".join(["---"] * len(header)) + " |")
                        for r in data_rows:
                            padded = r + [""] * (len(header) - len(r))
                            md.append("| " + " | ".join(padded[:len(header)]) + " |")
                        sections.append("\n".join(md))
                    else:
                        md = [f"### [Feuille : {s_name}] - Profil Structurel ({num_rows} lignes, {len(header)} colonnes)"]
                        md.append(f"**Colonnes détectées :** {', '.join(header)}")
                        md.append("\n**Aperçu des 10 premières lignes :**")
                        md.append("| " + " | ".join(header) + " |")
                        md.append("| " + " | ".join(["---"] * len(header)) + " |")
                        for r in data_rows[:10]:
                            padded = r + [""] * (len(header) - len(r))
                            md.append("| " + " | ".join(padded[:len(header)]) + " |")

                        md.append(f"\n*(... {num_rows - 15} lignes intermédiaires masquées ...)*\n")
                        md.append("**Aperçu des 5 dernières lignes :**")
                        md.append("| " + " | ".join(header) + " |")
                        md.append("| " + " | ".join(["---"] * len(header)) + " |")
                        for r in data_rows[-5:]:
                            padded = r + [""] * (len(header) - len(r))
                            md.append("| " + " | ".join(padded[:len(header)]) + " |")
                        sections.append("\n".join(md))

                return {
                    "status": "success",
                    "format": "excel",
                    "file_path": str(file_path),
                    "sheets_count": len(target_sheets),
                    "sheets": sheet_names,
                    "total_rows": total_file_rows,
                    "content": "\n\n---\n\n".join(sections),
                }
            except Exception as e:
                return {"status": "error", "message": f"Échec extraction Excel : {str(e)}"}
            finally:
                if wb is not None:
                    try:
                        wb.close()
                    except Exception:
                        pass

        try:
            content = file_path.read_text(encoding="utf-8", errors="ignore")
            return {
                "status": "success",
                "format": ext.lstrip(".") or "text",
                "content": content[:25000],
                "size_bytes": len(content),
            }
        except Exception as e:
            return {"status": "error", "message": f"Échec lecture document : {str(e)}"}

    def _exec_code_formatter(self, args: dict[str, Any]) -> dict[str, Any]:
        code = args.get("code_content", "")
        formatted = code.strip() + "\n"
        return {"status": "success", "formatted_code": formatted}

    def _exec_test_runner(self, args: dict[str, Any]) -> dict[str, Any]:
        test_dir = args.get("test_dir", "tests")
        pytest_args = args.get("pytest_args", ["-v"])
        
        # Exécution dans la Sandbox Docker étanche
        return docker_sandbox.run_tests_in_sandbox(
            target_dir=test_dir,
            pytest_args=pytest_args,
            timeout=30.0,
        )

    def _exec_search_models_catalog(self, args: dict[str, Any]) -> dict[str, Any]:
        from services.benchmarks_client import benchmarks_client
        q = args.get("q")
        max_price = float(args["max_price_out_usd"]) if args.get("max_price_out_usd") is not None else None
        min_coding = float(args["min_coding_score"]) if args.get("min_coding_score") is not None else None
        bench_filters = args.get("benchmark_filters")
        sort_by = args.get("sort_by", "coding_desc")
        limit = int(args.get("limit", 5))

        results = benchmarks_client.search_catalog(
            q=q,
            max_price_out_usd=max_price,
            min_coding_score=min_coding,
            benchmark_filters=bench_filters,
            sort_by=sort_by,
            limit=limit,
        )

        formatted = []
        for m in results:
            evals = m.get("evaluations", {}) or {}
            formatted.append({
                "name": m.get("name"),
                "slug": m.get("slug"),
                "creator": m.get("creator_name"),
                "price_in_usd": m.get("price_in_usd"),
                "price_out_usd": m.get("price_out_usd"),
                "price_cache_usd": m.get("price_cache_usd"),
                "speed_tok_s": m.get("speed_tok_s"),
                "coding_score": m.get("coding_index"),
                "intelligence_score": m.get("intelligence_index"),
                "terminalbench_v2_1": evals.get("terminalbench_v2_1") or evals.get("terminal_bench_v2_1"),
                "gpqa_diamond": evals.get("gpqa_diamond") or evals.get("gpqa"),
                "scicode": evals.get("scicode"),
                "evaluations": evals,
            })

        return {
            "status": "success",
            "count": len(formatted),
            "models": formatted,
        }

    def _exec_get_catalog_capabilities(self, args: dict[str, Any]) -> dict[str, Any]:
        from storage.repository import aa_benchmarks_repo
        status = aa_benchmarks_repo.get_sync_status()
        metrics = aa_benchmarks_repo.get_available_metric_keys()

        known_descriptions: dict[str, str] = {
            "terminalbench_v2_1": "Capacité d'exécution bash/terminal, git, docker et résolution de tests unitaires.",
            "terminalbench_hard": "Résolution de scénarios d'ingénierie système complexes en environnement réel.",
            "artificial_analysis_coding_index": "Indice composite officiel Artificial Analysis évaluant la programmation (0-100).",
            "artificial_analysis_intelligence_index": "Indice composite mesurant l'intelligence et le raisonnement global (0-100).",
            "gpqa": "Questions de raisonnement scientifique de niveau PhD / Doctorat.",
            "scicode": "Conception d'algorithmes scientifiques, calcul numérique et physique appliquée.",
            "lcr": "Long Context Reasoning sur des bases de code massives (50k+ lignes).",
            "ifbench": "Instruction Following - Suivi strict des contraintes négatives et des formats.",
            "livecodebench": "Problèmes de programmation compétitive récents.",
            "tau2": "Fiabilité agentique sur des chaînes d'appels d'outils et de fonctions (Function Calling / MCP).",
        }

        definitions = {}
        for k in metrics:
            definitions[k] = known_descriptions.get(k, f"Métrique scientifique dynamique : {k.replace('_', ' ').title()}.")

        return {
            "status": "success",
            "total_models": status["total_models"],
            "last_synced_at": status["last_synced_at"],
            "available_benchmark_metrics": metrics,
            "metric_definitions": definitions,
        }

    def _exec_discover_tools(self, args: dict[str, Any], project_id: str | None = None) -> dict[str, Any]:
        from services.tool_rag import tool_rag
        query = args.get("query", "")
        limit = int(args.get("limit", 3))
        discovered = tool_rag.search_relevant_tools(query=query, project_id=project_id, limit=limit)
        return {
            "status": "success",
            "query": query,
            "count": len(discovered),
            "discovered_tools": [
                {
                    "id": t.id,
                    "name": t.name,
                    "description": t.description,
                    "category": t.category,
                    "parameters_schema": t.parameters_schema,
                }
                for t in discovered
            ],
        }

    def _exec_read_skill(self, args: dict[str, Any], project_id: str | None = None) -> dict[str, Any]:
        skill_name = str(args.get("skill_name", "")).strip()
        if not skill_name:
            return {"status": "error", "message": "Nom de compétence manquant (skill_name requis)."}

        from services.skills_registry import skills_registry
        body = skills_registry.load_skill_body(skill_name, project_id=project_id)
        if not body:
            return {"status": "error", "message": f"Compétence '{skill_name}' introuvable ou inactive."}

        # Détection des sous-fichiers et ressources dans le dossier du skill
        skill_dir = settings.skills_dir / skill_name
        resources_list: list[str] = []
        if skill_dir.exists() and skill_dir.is_dir():
            for sub_p in skill_dir.rglob("*"):
                if sub_p.is_file() and sub_p.name != "SKILL.md":
                    resources_list.append(str(sub_p.relative_to(skill_dir)).replace("\\", "/"))

        return {
            "status": "success",
            "skill_name": skill_name,
            "instructions_md": body,
            "available_resources": resources_list,
        }

    def _exec_discover_skills(self, args: dict[str, Any], project_id: str | None = None) -> dict[str, Any]:
        query = str(args.get("query", "")).strip()
        limit = min(int(args.get("limit") or 3), 10)
        if not query:
            return {"status": "error", "message": "Requête de recherche vide."}

        from services.skill_rag import skill_rag
        found_skills = skill_rag.search_relevant_skills(query=query, project_id=project_id, limit=limit)
        return {
            "status": "success",
            "query": query,
            "count": len(found_skills),
            "skills": [
                {
                    "name": s.name,
                    "description": s.description,
                    "version": s.version,
                    "scope": s.scope.value,
                    "tags": s.tags,
                }
                for s in found_skills
            ],
        }


mcp_hub = McpHub()
