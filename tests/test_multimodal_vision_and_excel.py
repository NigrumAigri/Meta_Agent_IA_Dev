import openpyxl
import pytest
from pathlib import Path
from tempfile import NamedTemporaryFile
from fastapi.testclient import TestClient

from api.app import app
from services.mcp_hub import mcp_hub
from storage.repository import openrouter_models_repo, project_repo
from core.domain import Project, DocumentAttachment


@pytest.fixture
def client():
    return TestClient(app)


def test_excel_document_extractor_multi_sheets():
    """Vérifie l'ingestion générique d'un classeur Excel multi-feuilles avec mode adaptatif."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Synthese"
    ws1.append(["Projet", "Statut", "Budget_USD"])
    ws1.append(["Meta_Agent_V5", "En_Production", 150.0])
    ws1.append(["Sandbox_FinOps", "Actif", 50.0])

    ws2 = wb.create_sheet(title="BigData_Transactions")
    ws2.append(["ID", "Description", "Montant", "Date"])
    for i in range(1, 220):
        ws2.append([i, f"Tx_{i}", round(i * 12.5, 2), "2026-08-29"])

    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
        wb.save(tmp_path)
    wb.close()

    try:
        res = mcp_hub.execute_tool("document_extractor", {"file_path": str(tmp_path)})
        assert res["status"] == "success"
        assert res["format"] == "excel"
        assert "Synthese" in res["sheets"]
        assert "BigData_Transactions" in res["sheets"]
        assert res["total_rows"] == 223  # 3 + 220

        # Vérification du format Markdown propre
        content = res["content"]
        assert "### [Feuille : Synthese]" in content
        assert "| Projet | Statut | Budget_USD |" in content
        assert "### [Feuille : BigData_Transactions] - Profil Structurel" in content
        assert "Colonnes détectées" in content
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def test_excel_document_extractor_slicing():
    """Vérifie la capacité des agents d'extraire une tranche précise de lignes d'un onglet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Instruments"
    ws.append(["ISIN", "Devise", "Valeur"])
    for i in range(1, 100):
        ws.append([f"FR000{i:04d}", "EUR", i * 10])

    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
        wb.save(tmp_path)
    wb.close()

    try:
        res = mcp_hub.execute_tool(
            "document_extractor",
            {
                "file_path": str(tmp_path),
                "sheet_name": "Instruments",
                "start_row": 30,
                "max_rows": 5,
            },
        )
        assert res["status"] == "success"
        assert res["format"] == "excel"
        content = res["content"]
        assert "Extrait lignes 30 à 34" in content
        assert "FR0000029" in content
        assert "FR0000030" in content
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def test_vision_models_query_and_api(client: TestClient):
    """Vérifie que la découverte des modèles Vision dans SQLite est 100% dynamique et exposée par l'API."""
    # 1. Test Repository
    models = openrouter_models_repo.list_vision_models(limit=300)
    assert len(models) >= 50
    for m in models:
        assert m["is_vision"] is True
        assert "id" in m
        assert "pout" in m

    # 2. Test Endpoint API
    res = client.get("/api/v1/finops/models/vision?limit=20")
    assert res.status_code == 200
    data = res.json()
    assert isinstance(data, list)
    assert len(data) == 20
    assert all(m["is_vision"] is True for m in data)

    # 3. Test Filtre de recherche dynamique
    res_search = client.get("/api/v1/finops/models/vision?q=gemini")
    assert res_search.status_code == 200
    data_search = res_search.json()
    assert len(data_search) > 0
    assert all(
        "gemini" in m["id"].lower()
        or "gemini" in m["name"].lower()
        or "gemini" in (m.get("description") or "").lower()
        for m in data_search
    )


@pytest.mark.asyncio
async def test_multimodal_cadrage_turn_with_images_and_excel():
    """Vérifie que l'orchestrateur ingère des pièces jointes d'images et de tableurs sans crash."""
    from services.orchestrator import orchestrator

    test_proj = Project(name="Projet Test Multimodal Vision")
    project_repo.save(test_proj)

    img_att = DocumentAttachment(
        filename="maquette_ui.png",
        content_type="image/png",
        raw_content="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg==",
        size_bytes=100,
    )

    doc_att = DocumentAttachment(
        filename="specs_financieres.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        raw_content="### [Feuille : Données]\n| Code | Montant |\n| --- | --- |\n| A01 | 1500 |",
        size_bytes=50,
    )

    updated_proj = await orchestrator.run_cadrage_turn(
        project=test_proj,
        user_message="Analyse cette maquette et ces données financières pour cadrer le projet.",
        attachments=[img_att, doc_att],
    )

    thread = updated_proj.get_or_create_main_thread()
    assert len(thread.messages) >= 2
    user_msg = thread.messages[0]
    assert len(user_msg.attachments) == 2
    assert user_msg.attachments[0].filename == "maquette_ui.png"
    assert user_msg.attachments[1].filename == "specs_financieres.xlsx"

    # Nettoyage
    project_repo.delete(test_proj.id)
